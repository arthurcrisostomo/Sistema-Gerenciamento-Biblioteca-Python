
from openpyxl import load_workbook
import validacoes
import animacoes
from senha import Senha
import time

class Cadastrar():
    def __init__(self):
        # Arquivo excel com o banco de dados da biblioteca
        self.arquivo_excel = 'Dados Biblioteca.xlsx'
        # > Carrega o arquivo excel que contem o banco de dados da biblioteca
        self.wb = load_workbook(filename=self.arquivo_excel)

        # > Abre a planilha na aba 'Livros'
        self.ws = self.wb['Livros']

    def adicionar_livro(self, instancia_interface):
        """
        Essa função coleta o titulo, autor, ano,
        genero e disponibilidade do livro a ser 
        adicionado e adiciona no banco de dados da biblioteca
        caso não tenha nenhum erro, se não exibe o erro ao usuario
        e encerra
        """
        
        # > Instancia a classe Senha()
        senha = Senha()
        
        # > Exibe o modelo de senha
        instancia_interface.style_senha_adm()

        # > Atribui o valor booleano retornado a variavel acesso_liberado
        acesso_liberado = senha.verificar_senha()

        # > Se acesso liberador for false, não permite prosseguir e sai da função
        if not acesso_liberado:
            animacoes.loading_saindo()
            return

        while True:
            # > Define os inputs necessarios para adição de um livro
            dicionario_inputs = {
                'titulo': '\n\033[1;38;2;101;67;33mTitulo ⤷ \033[0m',
                'autor': '\n\033[1;38;2;101;67;33mAutor ⤷ \033[0m',
                'ano': '\n\033[1;38;2;101;67;33mAno ⤷ \033[0m',
                'genero': '\n\033[1;38;2;101;67;33mGenero ⤷ \033[0m'
            }

            # > Armazena os inputs na lista respostas
            respostas = []

            # > Itera sobre cada chave em dicionario_inputs
            for prompt in dicionario_inputs:

                # > Exibe o modelo de como sair do menu adicionar livro
                instancia_interface.style_aviso_adicionar_encerrar()

                # > Pega o valor da chave que esta passando e coloca em um input.title()
                propmt_input = input(dicionario_inputs[prompt]).title()

                # > Se a resposta for igual a "/" sai da função
                if propmt_input.strip() == '0':
                    animacoes.loading_saindo()
                    return

                erro = validacoes.validar_campos_cadastro(prompt, propmt_input)

                while erro:
                    instancia_interface.style_erro_cadastro(erro)
                    propmt_input = input(dicionario_inputs[prompt]).title()
                    
                    # > Se a resposta for igual a "/" sai da função
                    if propmt_input.strip() == '0':
                        animacoes.loading_saindo()
                        return
                    erro = validacoes.validar_campos_cadastro(prompt, propmt_input)

                # > Ao passar pelas verificacoes, armazena as respostas em respostas
                respostas.append(propmt_input)

            # > Ao final adiciona a Disponibilidade
            respostas.append('Disponível')

            # > Adiciona na planilha: Titulo | Autor | Ano | Genero | Disponibilidade
            self.ws.append(respostas)

            # > Salva o arquivo com as alterações
            self.wb.save(filename=self.arquivo_excel)

            # > Limpa a tela
            instancia_interface.limpar_tela()
            # > Inicia a animação de adição de livro
            animacoes.loading_adicionando_livro()

            # > Exibe o modelo de mensagem informado que o livro foi adicionado com sucesso
            instancia_interface.style_cadastro_concluido()

    def listar_livros(self, instancia_inteface):
        """
        Essa função percorre cada celula da coluna A + 1(Ate o ultimo item), ['A'] é apenas uma referencia
        para saber quantas linhas existem, acessa a worksheet(ws), na linha que esta passando começando de 2
        junta a letra do livro A + (numero da linha convertido para string) = A2, e pega o valor da A2 que
        é o titulo do livro usndo o .value, converte pra string e atribui a variavel, fazendo isso para todas
        as outras.
        """

        # > Verifica se não tem nenhum livro pra ser listado
        erro_livros_disponiveis = validacoes.validar_livros_listar(self.ws)
        
        if erro_livros_disponiveis:
            instancia_inteface.style_exibir_erro_listar(erro_livros_disponiveis)
            return

        livros = []

        for linha in range(2, len(self.ws['A']) + 1):
            livro = str(self.ws['A' + str(linha)].value)
            autor = str(self.ws['B' + str(linha)].value)
            ano = str(self.ws['C' + str(linha)].value)
            genero = str(self.ws['D' + str(linha)].value)
            disponibilidade = str(self.ws['E' + str(linha)].value)
            livros.append((livro, autor, ano, genero, disponibilidade))

        instancia_inteface.limpar_tela()
        instancia_inteface.style_exibir_livros(livros)
        user_opcao_exibicao = input('\033[38;2;222;184;135m⤷ 📚\033[0m')
        erro_exibicao_livros = validacoes.verificar_exibicao(user_opcao_exibicao)
        while erro_exibicao_livros:
            instancia_inteface.style_exibir_erro_listar(erro_exibicao_livros)
            instancia_inteface.style_exibir_livros(livros)
            user_opcao_exibicao = input('\033[38;2;222;184;135m⤷ 📚\033[0m')
            erro_exibicao_livros = validacoes.verificar_exibicao(user_opcao_exibicao)

        animacoes.loading_saindo()

    def buscar_livros(self, instancia_interface):
        instancia_interface.style_aviso_encerrar_busca()

        while True:

            # > Verifica se não tem nenhum livro pra ser listado
            erro_livros_disponiveis = validacoes.validar_campo_busca(self.ws)
        
            if erro_livros_disponiveis:
                instancia_interface.style_exibir_erro_busca(erro_livros_disponiveis)
                return

            busca_usuario = input('\033[38;2;222;184;135m⤷ 🔍\033[0m').title().strip()

            if busca_usuario.strip() == '0':
                animacoes.loading_saindo()
                return

            livros = []

            for linha in range(2, self.ws.max_row + 1):
                
                titulo = self.ws['A' + str(linha)].value
                if len(titulo) > 17:
                    titulo = titulo[:17] + '...'

                if busca_usuario in titulo:
                    autor = self.ws['B' + str(linha)].value
                    ano = self.ws['C' + str(linha)].value
                    genero = self.ws['D' + str(linha)].value
                    disponibilidade = self.ws['E' + str(linha)].value
                
                    livros.append((titulo, autor, ano, genero, disponibilidade))
            
            erro_busca = validacoes.verificar_busca_livros(busca_usuario, livros)
            if erro_busca:
                instancia_interface.style_exibir_erro_busca(erro_busca)
                instancia_interface.style_aviso_encerrar_busca()
            
            else:
                instancia_interface.style_exibir_livros(livros)


    def alugar_livro(self, instancia_interface):

        while True:
            # Livros disponiveis para ser alugado
            livros_disponiveis = []

            # Verifica o banco de dados e armazena em livros disponiveis
            # os livros disponiveis para alugar
            for linha in range(2, self.ws.max_row + 1):
                disponilidade_atual = self.ws[f'E{linha}'].value

                if disponilidade_atual == 'Disponível':
                    titulo = self.ws[f'A{linha}'].value
                    autor = self.ws[f'B{linha}'].value
                    ano = self.ws[f'C{linha}'].value
                    genero = self.ws[f'D{linha}'].value

                    livros_disponiveis.append((titulo, autor, ano, genero, disponilidade_atual))


            # > Verifica se não tem nenhum livro pra ser alugado
            erro_livros_disponiveis = validacoes.validar_livros(self.ws)

            # > Se tiver mensagem de erro, exibe a mensagem e volta pro menu
            if erro_livros_disponiveis:
                instancia_interface.style_erro_alugar(erro_livros_disponiveis)
                return

            # > Exibe os livros disponiveos para alugar
            instancia_interface.style_exibir_livros_disponiveis(livros_disponiveis)

            # > Input do usuario
            busca_usuario = input('\033[38;2;222;184;135m⤷ 🤚\033[0m').title().strip()

            # > Se digitar 0, aciona a animaçao de sair e volta pro menu
            if busca_usuario.strip() == '0':
                animacoes.loading_saindo()
                return
            
            # Verifica se o campo esta vazio ou se tem livros disponiveis para alugar
            erro_alugar_livro = validacoes.validar_campo_alugar(busca_usuario, self.ws)
            
            # Caso tenha mensagem de erro, se for erro de input vazio, apenas exibe e reseta o while
            # se for erro de livro indisponivel, exibe a mensagem e volta pro menu
            if erro_alugar_livro:
                if 'Vazio' in erro_alugar_livro:
                    instancia_interface.style_erro_alugar(erro_alugar_livro)
                else:
                    instancia_interface.style_erro_alugar(erro_alugar_livro)
                
            # Caso não tenha mensagem de erro prossegue no processo de alugar livro
            else:
                for linha in range(2, self.ws.max_row + 1):
                    titulo = self.ws[f'A{linha}'].value

                    if busca_usuario.title().strip() == titulo:
                        self.ws[f'E{linha}'] = 'Alugado'
                        self.wb.save(self.arquivo_excel)

                animacoes.loading_alugando_livro()
                instancia_interface.style_livro_alugado_concluido()

    def devolver_livro(self, instancia_interface):

        while True:
            instancia_interface.style_aviso_encerrar_devolver()

            # > Verifica se tem algum livro com disponibilidade como Alugado
            # caso tenha, imprime a mensagem e volta pro menu
            erro_devolver = validacoes.validar_livros_alugados(self.ws)
            if erro_devolver:
                instancia_interface.style_erro_devolver(erro_devolver)
                return

            # > Armazena todos os livros alugados atuais do banco de dados
            livros_alugados = []

            for linha in range(2, self.ws.max_row + 1):
                titulo = self.ws[f'A{linha}'].value
                disponibilidade = self.ws[f'E{linha}'].value

                if disponibilidade == 'Alugado':
                    autor = self.ws[f'B{linha}'].value
                    ano = self.ws[f'C{linha}'].value
                    genero = self.ws[f'D{linha}'].value
                    livros_alugados.append((titulo, autor, ano, genero, disponibilidade))
            
            # > Exibe todos os livros com disponibilidade alugado
            instancia_interface.exibir_livros(livros_alugados)

            # > Input usuario
            livro_devolver_usuario = input('\033[38;2;222;184;135m⤷ 🗳️ \033[0m').title().strip()
            
            for linha in range(2, self.ws.max_row + 1):
                titulo = self.ws[f'A{linha}'].value

                if livro_devolver_usuario == titulo:
                    self.ws[f'E{linha}'].value = 'Disponível'
                    self.wb.save(self.arquivo_excel)
                    animacoes.loading_devolvendo_livro()
                    instancia_interface.style_livro_devolvido_concluido()
    
    def remover_livro(self, instancia_interface):
        # > Instancia a classe Senha()
        senha = Senha()
        
        # > Exibe o modelo de senha
        instancia_interface.style_senha_adm()

        # > Atribui o valor booleano retornado a variavel acesso_liberado
        acesso_liberado = senha.verificar_senha()

        # > Se acesso liberador for false, não permite prosseguir e sai da função
        if not acesso_liberado:
            animacoes.loading_saindo()
            return
        
        while True:
            instancia_interface.style_aviso_encerrar_remover()

            erro_livro_disponivel_remover = validacoes.validar_livros_remover(self.ws)
            if erro_livro_disponivel_remover:
                instancia_interface.style_erro_remover(erro_livro_disponivel_remover)
                return
            
            livros = []

            for linha in range(2, self.ws.max_row + 1):
                titulo = self.ws[f'A{linha}'].value
                autor = self.ws[f'B{linha}'].value
                ano = self.ws[f'C{linha}'].value
                genero = self.ws[f'D{linha}'].value
                disponibilidade = self.ws[f'D{linha}'].value

                livros.append((titulo, autor, ano, genero, disponibilidade))

            instancia_interface.exibir_livros(livros)

            livro_remover_usuario = input('\033[38;2;222;184;135m⤷ 🚫\033[0m').title().strip()

            erro_remover_vazio = validacoes.validar_campo_remocao(livro_remover_usuario, self.ws)

            if erro_remover_vazio:
                instancia_interface.style_aviso_encerrar_remover()
                instancia_interface.style_erro_remover(erro_remover_vazio)
            
            else:
                for linha in range(2, self.ws.max_row + 1):
                    titulo = self.ws[f'A{linha}'].value
                
                    if livro_remover_usuario == titulo:
                        self.ws.delete_rows(linha)
                        self.wb.save(self.arquivo_excel)

                        animacoes.loading_removendo_livro()
                        instancia_interface.style_livro_removido_sucesso()

            

            

            
                
